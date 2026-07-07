"""
Endpoints para Entrada/Salida (registro diario, reemplaza Excel).
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query, Body, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from datetime import date, datetime
import re
from app.database import get_db
from app.models import EntradaSalida, Expediente, Usuario, Notificacion, BorradoListado
from app.schemas import (
    EntradaSalida as EntradaSalidaSchema, EntradaSalidaCreate,
    BorradoListado as BorradoListadoSchema,
)
from app.utils.deps import obtener_usuario_actual

router = APIRouter(prefix="/api/entrada-salida", tags=["entrada-salida"])


@router.post("/", response_model=EntradaSalidaSchema)
async def crear_entrada_salida(
    entrada_create: EntradaSalidaCreate,
    db: Session = Depends(get_db)
):
    """
    Crea un registro de Entrada/Salida.
    Puede estar vinculado a un expediente existente o crear uno nuevo.
    """
    # Si viene número de expediente, asociarlo (creando el expediente si no existe)
    expediente_id = None
    if entrada_create.numero_expediente:
        numero = entrada_create.numero_expediente.replace("*", "").strip()
        expediente = db.query(Expediente).filter(Expediente.numero == numero).first()
        if not expediente:
            # Crear el registro único del expediente con los datos del listado
            despachante = None
            if entrada_create.asignacion:
                despachante = db.query(Usuario).filter(
                    Usuario.nombre == entrada_create.asignacion.strip()
                ).first()
            expediente = Expediente(
                numero=numero,
                juzgado=entrada_create.juzgado,
                caratula=entrada_create.autos,
                estado="activo",
                despachante_id=despachante.id if despachante else None,
                fecha_entrada=entrada_create.fecha,
                conexos=[],
                observaciones=entrada_create.observaciones or "",
            )
            db.add(expediente)
            db.flush()
        expediente_id = expediente.id

    nueva_entrada = EntradaSalida(
        fecha=entrada_create.fecha,
        juzgado=entrada_create.juzgado,
        expediente_id=expediente_id,
        autos=entrada_create.autos,
        asignacion=entrada_create.asignacion,
        pase_firma=entrada_create.pase_firma,
        subido_lex=entrada_create.subido_lex,
        observaciones=entrada_create.observaciones,
        subido_defensa=entrada_create.subido_defensa,
        urgente=entrada_create.urgente,
    )

    db.add(nueva_entrada)

    # Si es urgente y está asignado a alguien, notificarle en su pantalla de inicio
    if entrada_create.urgente and entrada_create.asignacion:
        asignado = db.query(Usuario).filter(
            Usuario.nombre == entrada_create.asignacion.strip()
        ).first()
        if asignado:
            db.add(Notificacion(
                usuario_id=asignado.id,
                tipo="expediente_urgente",
                contenido=f"Se te asignó un expediente URGENTE: {entrada_create.numero_expediente or ''} — {entrada_create.autos[:80]}",
                expediente_id=expediente_id,
            ))

    db.commit()
    db.refresh(nueva_entrada)

    return nueva_entrada


def _insertar_filas(db, filas):
    """
    Carga filas del listado pensada para re-pegar el Excel completo cada día:
      - Fila NUEVA (no coincide con nada) → se agrega.
      - Fila que YA está (misma fecha + expediente + juzgado + autos + asignación):
        si el Excel trae fecha de "pase a la firma" o "subido al Lex" que el
        sistema no tenía (o cambió), se ACTUALIZA esa fila (y se repinta sola).
        Si no hay nada nuevo, se saltea.
    Devuelve {creados, actualizados, omitidos}.
    """
    import unicodedata

    def _norm(s):
        # minúsculas, sin tildes, sin espacios de más → tolera diferencias mínimas
        s = " ".join((s or "").strip().split()).lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

    def _firma(fecha, numero, juzgado, autos, asignacion):
        return (fecha.isoformat() if fecha else "", _norm(numero), _norm(juzgado), _norm(autos), _norm(asignacion))

    # Mapa firma → fila existente (la primera que coincida).
    existentes = {}
    for fila_db, nu in (
        db.query(EntradaSalida, Expediente.numero)
        .outerjoin(Expediente, EntradaSalida.expediente_id == Expediente.id)
        .all()
    ):
        existentes.setdefault(_firma(fila_db.fecha, nu, fila_db.juzgado, fila_db.autos, fila_db.asignacion), fila_db)

    creados = 0
    actualizados = 0
    omitidos = 0
    for f in filas:
        autos = (f.get("autos") or "").strip()
        numero = (f.get("numero_expediente") or "").replace("*", "").strip()
        if not autos and not numero:
            continue  # fila vacía
        fecha = _a_fecha(f.get("fecha")) or date.today()
        asignacion = (f.get("asignacion") or "").strip() or None

        firma = _firma(fecha, numero, f.get("juzgado"), autos, asignacion)
        if firma in existentes:
            # Ya estaba: ver si el Excel trae fechas nuevas para actualizar.
            fila_db = existentes[firma]
            pf = _a_fecha(f.get("pase_firma"))
            sl = _a_fecha(f.get("subido_lex"))
            cambio = False
            if pf and pf != fila_db.pase_firma:
                fila_db.pase_firma = pf
                cambio = True
            if sl and sl != fila_db.subido_lex:
                fila_db.subido_lex = sl
                cambio = True
            if cambio:
                actualizados += 1
            else:
                omitidos += 1
            continue

        expediente_id = None
        if numero:
            exp = db.query(Expediente).filter(Expediente.numero == numero).first()
            if not exp:
                desp = db.query(Usuario).filter(Usuario.nombre == asignacion).first() if asignacion else None
                exp = Expediente(
                    numero=numero, juzgado=f.get("juzgado"), caratula=autos, estado="activo",
                    despachante_id=desp.id if desp else None, fecha_entrada=fecha,
                    conexos=[], observaciones=f.get("observaciones") or "",
                )
                db.add(exp)
                db.flush()
            expediente_id = exp.id

        nueva = EntradaSalida(
            fecha=fecha, juzgado=f.get("juzgado"), expediente_id=expediente_id, autos=autos,
            asignacion=asignacion, pase_firma=_a_fecha(f.get("pase_firma")),
            subido_lex=_a_fecha(f.get("subido_lex")), observaciones=f.get("observaciones"),
            urgente=bool(f.get("urgente")),
        )
        db.add(nueva)
        existentes[firma] = nueva  # por si la misma fila viene repetida en el archivo
        creados += 1

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}


@router.post("/bulk")
async def crear_bulk(
    filas: list[dict] = Body(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual),
):
    """Crea varias filas del listado (pegado desde Excel), sin duplicar; actualiza fechas de firma/subido."""
    return _insertar_filas(db, filas)


def _campo_de_encabezado(h: str) -> str:
    """Traduce el nombre de una columna del Excel al campo interno."""
    x = (h or "").strip().lower()
    if "juzg" in x:
        return "juzgado"
    if "exped" in x or "n°" in x or "nro" in x:
        return "numero_expediente"
    if "auto" in x or "carat" in x or "carát" in x:
        return "autos"
    if "asign" in x or "despach" in x:
        return "asignacion"
    if "firma" in x or "pase" in x:
        return "pase_firma"
    if "lex" in x:
        return "subido_lex"
    if "observ" in x:
        return "observaciones"
    if "fecha" in x:
        return "fecha"
    return ""


@router.post("/importar-xlsx")
async def importar_xlsx(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual),
):
    """
    Importa un Excel (.xlsx) al listado. Usa la primera fila como encabezado
    para saber qué columna es cada dato, y NO duplica lo que ya está cargado.
    """
    from openpyxl import load_workbook
    import io

    if not (archivo.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Subí un archivo Excel (.xlsx).")
    contenido = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el Excel. ¿Es un .xlsx válido?")

    ws = wb.active
    filas_raw = list(ws.iter_rows(values_only=True))
    if not filas_raw:
        return {"creados": 0, "omitidos": 0}

    encabezados = [str(c) if c is not None else "" for c in filas_raw[0]]
    mapa = [_campo_de_encabezado(h) for h in encabezados]
    # Si no se reconoció ninguna columna, usar el orden del export del sistema.
    if not any(mapa):
        orden = ["fecha", "juzgado", "numero_expediente", "autos", "asignacion", "pase_firma", "subido_lex", "observaciones"]
        mapa = [orden[i] if i < len(orden) else "" for i in range(len(encabezados))]

    filas = []
    for row in filas_raw[1:]:
        obj = {}
        for i, campo in enumerate(mapa):
            if campo and i < len(row) and row[i] is not None:
                v = row[i]
                obj[campo] = v.strip() if isinstance(v, str) else str(v).strip()
        if obj.get("autos") or obj.get("numero_expediente"):
            filas.append(obj)

    return _insertar_filas(db, filas)


@router.get("/", response_model=list[EntradaSalidaSchema])
async def listar_entrada_salida(
    fecha_inicio: date = Query(None),
    fecha_fin: date = Query(None),
    juzgado: str = Query(None),
    asignacion: str = Query(None),
    busqueda: str = Query(None),
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db)
):
    """
    Lista el listado diario (Entrada/Salida) con filtros opcionales.
    Es la pantalla principal compartida.
    """
    query = _aplicar_filtros(
        db.query(EntradaSalida), fecha_inicio, fecha_fin, juzgado, asignacion, busqueda
    )
    registros = (
        query.order_by(EntradaSalida.fecha.desc(), EntradaSalida.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return registros


@router.get("/total")
async def total_entrada_salida(
    fecha_inicio: date = Query(None),
    fecha_fin: date = Query(None),
    juzgado: str = Query(None),
    asignacion: str = Query(None),
    busqueda: str = Query(None),
    db: Session = Depends(get_db),
):
    """Cantidad total de registros que cumplen los filtros (para la paginación)."""
    query = _aplicar_filtros(
        db.query(func.count(EntradaSalida.id)), fecha_inicio, fecha_fin, juzgado, asignacion, busqueda
    )
    return {"total": query.scalar()}


def _aplicar_filtros(query, fecha_inicio, fecha_fin, juzgado, asignacion, busqueda):
    """Aplica los filtros comunes a una query de EntradaSalida."""
    if fecha_inicio:
        query = query.filter(EntradaSalida.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(EntradaSalida.fecha <= fecha_fin)
    if juzgado:
        query = query.filter(EntradaSalida.juzgado == juzgado)
    if asignacion:
        query = query.filter(EntradaSalida.asignacion == asignacion)
    if busqueda:
        like = f"%{busqueda}%"
        query = query.outerjoin(Expediente, EntradaSalida.expediente_id == Expediente.id).filter(or_(
            EntradaSalida.autos.ilike(like),
            EntradaSalida.observaciones.ilike(like),
            EntradaSalida.juzgado.ilike(like),
            Expediente.numero.ilike(like),
        ))
    return query


@router.get("/borrados", response_model=list[BorradoListadoSchema])
async def listar_borrados(db: Session = Depends(get_db)):
    """Papelera: filas del listado que se borraron (las más recientes primero)."""
    return (
        db.query(BorradoListado)
        .order_by(BorradoListado.fecha_borrado.desc())
        .limit(300)
        .all()
    )


@router.get("/{entrada_id}", response_model=EntradaSalidaSchema)
async def obtener_entrada_salida(entrada_id: int, db: Session = Depends(get_db)):
    """
    Obtiene un registro específico.
    """
    entrada = db.query(EntradaSalida).filter(EntradaSalida.id == entrada_id).first()

    if not entrada:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Registro no encontrado"
        )

    return entrada


# Campos editables inline desde el listado (tipo Excel)
_EDITABLES = {"fecha", "juzgado", "autos", "asignacion", "pase_firma",
              "subido_lex", "observaciones", "urgente", "subido_defensa", "cancelada"}
_CAMPOS_FECHA = {"fecha", "pase_firma", "subido_lex"}


def _a_fecha(valor):
    """
    '' o None → None; date/datetime → date.
    Entiende 'aaaa-mm-dd' (ISO) y también 'dd/mm/aaaa' o 'dd-mm-aaaa'
    (como pega Excel), con o sin hora al final.
    """
    if valor in (None, "", "null"):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    s = str(valor).strip()
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
    if m:
        d_, m_, y_ = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y_ < 100:
            y_ += 2000
        try:
            return date(y_, m_, d_)
        except ValueError:
            return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


@router.put("/{entrada_id}", response_model=EntradaSalidaSchema)
async def actualizar_entrada_salida(
    entrada_id: int,
    entrada_update: dict,
    db: Session = Depends(get_db)
):
    """
    Actualiza un registro del listado (edición inline tipo Excel).
    Acepta cualquiera de los campos editables; las fechas se pueden vaciar (None).
    """
    entrada = db.query(EntradaSalida).filter(EntradaSalida.id == entrada_id).first()
    if not entrada:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro no encontrado")

    for key, value in entrada_update.items():
        if key not in _EDITABLES:
            continue
        if key in _CAMPOS_FECHA:
            value = _a_fecha(value)
        setattr(entrada, key, value)

    # Si se editaron las observaciones y hay un "conexos:", se suman solos al
    # legajo de la persona (cuando el expediente ya tiene legajo).
    if "observaciones" in entrada_update and entrada.expediente_id:
        from app.services import legajos as legajos_svc
        exp = db.query(Expediente).filter(Expediente.id == entrada.expediente_id).first()
        legajos_svc.capturar_desde_observaciones(db, exp, entrada.observaciones)

    db.commit()
    db.refresh(entrada)
    return entrada


@router.delete("/{entrada_id}")
async def eliminar_entrada_salida(
    entrada_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obtener_usuario_actual),
):
    """
    Elimina una fila del listado, guardando una copia en la papelera
    (borrados_listado) por si hace falta revisarla después.
    """
    entrada = db.query(EntradaSalida).filter(EntradaSalida.id == entrada_id).first()
    if not entrada:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro no encontrado")

    # Capturar los datos antes de tocar la sesión
    snap = {
        "fecha": entrada.fecha,
        "juzgado": entrada.juzgado,
        "numero_expediente": entrada.numero_expediente,
        "autos": entrada.autos,
        "asignacion": entrada.asignacion,
        "observaciones": entrada.observaciones,
        "borrado_por": usuario.nombre,
    }

    # Guardar copia en la papelera (best-effort: si falla, igual se borra la fila)
    try:
        db.add(BorradoListado(**snap))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[!] No se pudo registrar en la papelera: {e}")

    # Borrar la fila
    entrada = db.query(EntradaSalida).filter(EntradaSalida.id == entrada_id).first()
    if entrada:
        from app.utils.auditoria import registrar
        registrar(db, usuario, "borró", "listado", f"{snap.get('numero_expediente') or ''} {snap.get('autos') or ''}".strip())
        db.delete(entrada)
        db.commit()
    return {"message": "Registro eliminado"}


@router.post("/export/excel")
async def exportar_excel(
    fecha_inicio: date = Query(None),
    fecha_fin: date = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporta registros de Entrada/Salida a Excel.
    Retorna un archivo descargable.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    # Buscar registros
    query = db.query(EntradaSalida)

    if fecha_inicio:
        query = query.filter(EntradaSalida.fecha >= fecha_inicio)

    if fecha_fin:
        query = query.filter(EntradaSalida.fecha <= fecha_fin)

    registros = query.order_by(EntradaSalida.fecha.asc()).all()

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Entrada/Salida"

    # Encabezados
    headers = [
        "Fecha", "Juzgado", "Expediente", "Autos",
        "Asignación", "Pase a la firma", "Subido al Lex",
        "Observaciones", "Subido al Defensa"
    ]

    header_fill = PatternFill(start_color="1B2B42", end_color="1B2B42", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for row, reg in enumerate(registros, 2):
        ws.cell(row, 1).value = reg.fecha
        ws.cell(row, 2).value = reg.juzgado
        ws.cell(row, 3).value = reg.expediente.numero if reg.expediente else ""
        ws.cell(row, 4).value = reg.autos
        ws.cell(row, 5).value = reg.asignacion
        ws.cell(row, 6).value = reg.pase_firma
        ws.cell(row, 7).value = reg.subido_lex
        ws.cell(row, 8).value = reg.observaciones
        ws.cell(row, 9).value = "Sí" if reg.subido_defensa else "No"

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 15
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return {
        "filename": "entrada_salida.xlsx",
        "data": output.getvalue().hex()  # Convertir a hex para enviar en JSON
    }
